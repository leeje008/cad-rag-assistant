"""Document parsers producing section/page-aware markdown blocks.

PDF: pymupdf4llm (markdown per page)
XLSX: openpyxl + pandas (markdown per sheet, merged cells forward-filled)

Both yield `ParsedDoc` with enough metadata for the chunker to keep page
and section information through the retrieval pipeline.
"""

from __future__ import annotations

import base64
import hashlib
import io
import logging
import unicodedata
from dataclasses import dataclass, field
from pathlib import Path

from .settings import settings

logger = logging.getLogger(__name__)


@dataclass
class ParsedBlock:
    text: str
    page: int | None = None
    section: str | None = None
    block_type: str = "text"  # "text" | "table" | "image"
    bbox: tuple[float, float, float, float] | None = None
    table_id: str | None = None
    figure_id: str | None = None
    table_html: str | None = None  # original-preserving HTML for table blocks
    image_b64: str | None = None  # PNG base64 for image blocks (ingest-only, not stored)


@dataclass
class ParsedDoc:
    doc_id: str
    source_path: str
    title: str
    blocks: list[ParsedBlock] = field(default_factory=list)


def _normalise_path(path: Path) -> Path:
    return Path(unicodedata.normalize("NFC", str(path)))


def _doc_id(path: Path) -> str:
    path = _normalise_path(path)
    stat = path.stat()
    key = f"{path}|{stat.st_size}|{int(stat.st_mtime)}"
    return hashlib.sha1(key.encode("utf-8")).hexdigest()[:16]


def parse_pdf(path: Path) -> ParsedDoc:
    """Parse a PDF, preferring Docling (structure/table/image aware).

    Falls back to pymupdf4llm if Docling is disabled or raises — the router
    keeps the system resilient while Docling matures.
    """

    if settings.use_docling:
        try:
            return parse_pdf_docling(path)
        except Exception as exc:  # noqa: BLE001
            logger.warning(
                "docling failed for %s, falling back to pymupdf4llm: %s",
                Path(path).name,
                exc,
            )
    return parse_pdf_pymupdf(path)


_DOCLING_CONVERTER = None


def _docling_converter():
    """Lazily build and cache the Docling converter (loads models once)."""

    global _DOCLING_CONVERTER
    if _DOCLING_CONVERTER is None:
        from docling.datamodel.base_models import InputFormat
        from docling.datamodel.pipeline_options import PdfPipelineOptions
        from docling.document_converter import DocumentConverter, PdfFormatOption

        opts = PdfPipelineOptions()
        opts.generate_picture_images = True
        opts.images_scale = settings.docling_images_scale
        opts.do_ocr = settings.docling_do_ocr
        _DOCLING_CONVERTER = DocumentConverter(
            format_options={InputFormat.PDF: PdfFormatOption(pipeline_options=opts)}
        )
    return _DOCLING_CONVERTER


def _docling_prov(item, ddoc) -> tuple[int | None, tuple[float, float, float, float] | None]:
    """Extract (page, top-left-origin bbox) from a Docling item's provenance."""

    prov = getattr(item, "prov", None)
    if not prov:
        return None, None
    p = prov[0]
    page = _coerce_page(getattr(p, "page_no", None))
    bbox: tuple[float, float, float, float] | None = None
    try:
        bb = p.bbox
        page_obj = ddoc.pages[p.page_no]
        tl = bb.to_top_left_origin(page_height=page_obj.size.height)
        bbox = (float(tl.l), float(tl.t), float(tl.r), float(tl.b))
    except Exception:  # noqa: BLE001
        bbox = None
    return page, bbox


def _picture_b64(item, ddoc) -> str | None:
    """Render a Docling PictureItem to PNG base64 (requires image generation)."""

    try:
        img = item.get_image(ddoc)
        if img is None:
            return None
        buf = io.BytesIO()
        img.save(buf, format="PNG")
        return base64.b64encode(buf.getvalue()).decode("ascii")
    except Exception:  # noqa: BLE001
        return None


def parse_pdf_docling(path: Path) -> ParsedDoc:
    """Parse a PDF into typed blocks (text/table/image) with bbox via Docling."""

    from docling_core.types.doc.document import (
        PictureItem,
        SectionHeaderItem,
        TableItem,
        TextItem,
    )

    path = _normalise_path(path)
    doc_id = _doc_id(path)
    result = _docling_converter().convert(str(path))
    ddoc = result.document

    blocks: list[ParsedBlock] = []
    current_section: str | None = None
    tbl_i = 0
    fig_i = 0

    for item, _level in ddoc.iterate_items():
        # SectionHeaderItem subclasses TextItem — check it first.
        if isinstance(item, SectionHeaderItem):
            heading = (item.text or "").strip()
            if not heading:
                continue
            current_section = heading
            page, bbox = _docling_prov(item, ddoc)
            blocks.append(
                ParsedBlock(
                    text=f"# {heading}",
                    page=page,
                    section=current_section,
                    block_type="text",
                    bbox=bbox,
                )
            )
        elif isinstance(item, TableItem):
            md = item.export_to_markdown(doc=ddoc).strip()
            if not md:
                continue
            page, bbox = _docling_prov(item, ddoc)
            blocks.append(
                ParsedBlock(
                    text=md,
                    page=page,
                    section=current_section,
                    block_type="table",
                    bbox=bbox,
                    table_id=f"{doc_id}:tbl:{tbl_i}",
                    table_html=item.export_to_html(doc=ddoc),
                )
            )
            tbl_i += 1
        elif isinstance(item, PictureItem):
            b64 = _picture_b64(item, ddoc)
            if b64 is None:
                continue
            page, bbox = _docling_prov(item, ddoc)
            blocks.append(
                ParsedBlock(
                    text="",  # filled later by the VLM captioner (PR6)
                    page=page,
                    section=current_section,
                    block_type="image",
                    bbox=bbox,
                    figure_id=f"{doc_id}:fig:{fig_i}",
                    image_b64=b64,
                )
            )
            fig_i += 1
        elif isinstance(item, TextItem):
            text = (item.text or "").strip()
            if not text:
                continue
            page, bbox = _docling_prov(item, ddoc)
            blocks.append(
                ParsedBlock(
                    text=text,
                    page=page,
                    section=current_section,
                    block_type="text",
                    bbox=bbox,
                )
            )

    logger.info(
        "docling parsed %s → %d blocks (%d tables, %d images)",
        path.name,
        len(blocks),
        tbl_i,
        fig_i,
    )
    return ParsedDoc(
        doc_id=doc_id,
        source_path=str(path),
        title=path.name,
        blocks=blocks,
    )


def parse_pdf_pymupdf(path: Path) -> ParsedDoc:
    """Parse a PDF into one ParsedBlock per page using pymupdf4llm (fallback)."""

    import pymupdf4llm

    path = _normalise_path(path)
    raw = pymupdf4llm.to_markdown(str(path), page_chunks=True)

    blocks: list[ParsedBlock] = []
    for item in raw:
        # Newer versions return dicts; older tuple builds are unlikely but defended.
        if isinstance(item, dict):
            text = item.get("text", "") or ""
            meta = item.get("metadata") or {}
            page = meta.get("page")
            if page is None:
                page = item.get("page")
        elif isinstance(item, tuple) and item:
            text = item[0] if isinstance(item[0], str) else ""
            page = item[1] if len(item) > 1 and isinstance(item[1], int) else None
        else:
            continue

        text = (text or "").strip()
        if not text:
            continue

        section = _extract_first_heading(text)
        blocks.append(
            ParsedBlock(text=text, page=_coerce_page(page), section=section)
        )

    return ParsedDoc(
        doc_id=_doc_id(path),
        source_path=str(path),
        title=path.name,
        blocks=blocks,
    )


def parse_xlsx(path: Path) -> ParsedDoc:
    """Parse an XLSX workbook into one ParsedBlock per sheet."""

    import openpyxl
    import pandas as pd

    path = _normalise_path(path)
    doc_id = _doc_id(path)
    wb = openpyxl.load_workbook(str(path), data_only=True, read_only=False)

    blocks: list[ParsedBlock] = []
    for sheet_index, sheet_name in enumerate(wb.sheetnames, start=1):
        ws = wb[sheet_name]

        # Forward-fill merged cells so each logical row sees the header.
        for merged in list(ws.merged_cells.ranges):
            min_col, min_row, max_col, max_row = merged.bounds
            top_value = ws.cell(row=min_row, column=min_col).value
            ws.unmerge_cells(str(merged))
            for row in range(min_row, max_row + 1):
                for col in range(min_col, max_col + 1):
                    ws.cell(row=row, column=col).value = top_value

        rows = list(ws.iter_rows(values_only=True))
        if not rows:
            continue

        # Drop fully-empty leading rows.
        while rows and all(cell is None for cell in rows[0]):
            rows = rows[1:]
        if not rows:
            continue

        header = [
            str(cell) if cell is not None else f"col_{i}"
            for i, cell in enumerate(rows[0])
        ]
        data_rows = rows[1:]
        df = pd.DataFrame(data_rows, columns=header)
        if df.empty:
            continue

        clean = df.fillna("").astype(str)
        md_table = clean.to_markdown(index=False)
        block_text = f"# {sheet_name}\n\n{md_table}"
        blocks.append(
            ParsedBlock(
                text=block_text,
                page=sheet_index,
                section=sheet_name,
                block_type="table",
                table_id=f"{doc_id}:sheet:{sheet_index}",
                table_html=clean.to_html(index=False),
            )
        )

    wb.close()
    return ParsedDoc(
        doc_id=doc_id,
        source_path=str(path),
        title=path.name,
        blocks=blocks,
    )


def parse_any(path: Path) -> ParsedDoc:
    path = _normalise_path(path)
    suffix = path.suffix.lower()
    if suffix == ".pdf":
        return parse_pdf(path)
    if suffix in {".xlsx", ".xlsm"}:
        return parse_xlsx(path)
    raise ValueError(f"Unsupported file type: {suffix} ({path.name})")


# ---------------------------------------------------------------------------


def _coerce_page(page: object) -> int | None:
    if page is None:
        return None
    try:
        return int(page)
    except (TypeError, ValueError):
        return None


def _extract_first_heading(text: str) -> str | None:
    for raw_line in text.splitlines():
        line = raw_line.strip()
        if line.startswith("#"):
            return line.lstrip("#").strip() or None
    return None


SUPPORTED_SUFFIXES = {".pdf", ".xlsx", ".xlsm"}
